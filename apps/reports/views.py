"""
Reports - Sales analytics, exportable reports
"""
from django.shortcuts import render
from django.contrib.auth.decorators import login_required
from django.http import HttpResponse
from django.db.models import Sum, Count
from django.db.models.functions import TruncDate
from django.utils import timezone
from datetime import timedelta, date as date_cls
from apps.orders.models import Order, OrderItem
from apps.accounts.decorators import admin_required


# ── Shared helpers ────────────────────────────────────────────────────────────

def _parse_date_range(request, today):
    """
    Parse and validate the ``start`` / ``end`` GET parameters.

    Returns a ``(start_date_str, end_date_str)`` tuple of ISO-format date
    strings ready for use in ORM ``__date__gte`` / ``__date__lte`` filters.

    Validation rules (both views share identical logic):
    - Invalid or missing values fall back to the default 30-day window.
    - An inverted range (start > end) is silently swapped.

    Extracted to eliminate the verbatim duplication between
    ``reports_index`` and ``export_excel``.
    """
    raw_start = request.GET.get('start', '')
    raw_end   = request.GET.get('end',   '')
    try:
        start_date = date_cls.fromisoformat(raw_start).isoformat()
    except (ValueError, TypeError):
        start_date = (today - timedelta(days=30)).isoformat()
    try:
        end_date = date_cls.fromisoformat(raw_end).isoformat()
    except (ValueError, TypeError):
        end_date = today.isoformat()
    # Guard against inverted ranges (start after end) — swap silently.
    if start_date > end_date:
        start_date, end_date = end_date, start_date
    return start_date, end_date


def _base_orders_qs(start_date, end_date):
    """
    Return the base ``Order`` queryset shared by both report views.

    Applies the authoritative sales filter — ``is_paid=True`` AND
    ``status='completed'`` — that is consistent with Finance and Dashboard.
    Only the date range varies between calls; the eligibility criteria are
    defined once here so neither view can silently diverge from the other.

    ``reports_index``  uses the queryset directly for aggregations.
    ``export_excel``   chains ``.select_related('cashier').annotate(...)``
                       on top for the per-row Excel output.
    """
    return Order.objects.filter(
        is_paid=True,
        status='completed',
        created_at__date__gte=start_date,
        created_at__date__lte=end_date,
    )


# ── Views ─────────────────────────────────────────────────────────────────────

@login_required
@admin_required
def reports_index(request):
    today = timezone.localdate()
    start_date, end_date = _parse_date_range(request, today)
    orders = _base_orders_qs(start_date, end_date)

    # OPT-1: single aggregate query returns both total_revenue and total_orders
    # in one DB round-trip instead of two (aggregate then count separately).
    agg = orders.aggregate(total=Sum('total'), count=Count('id'))
    total_revenue = agg['total'] or 0
    total_orders  = agg['count'] or 0
    avg_order = total_revenue / total_orders if total_orders > 0 else 0

    # Sales by day — ONE grouped query for the whole range instead of two
    # queries per day (N+1). Days without sales are filled in below so the
    # chart keeps its zero entries.
    daily_sales = []
    start = date_cls.fromisoformat(start_date)
    end   = date_cls.fromisoformat(end_date)
    delta = end - start
    day_rows = {
        row['day']: row
        for row in orders.annotate(day=TruncDate('created_at'))
        .values('day')
        .annotate(total=Sum('total'), count=Count('id'))
    }
    for i in range(delta.days + 1):
        day = start + timedelta(days=i)
        row = day_rows.get(day) or {}
        daily_sales.append({
            'date':  day.strftime('%b %d'),
            'total': float(row.get('total') or 0),
            'count': row.get('count') or 0,
        })

    # Top products — ranked by revenue (existing intended behaviour).
    # Secondary sort on product_name makes ties alphabetically deterministic
    # so the same dataset always produces the same report order.
    top_products = OrderItem.objects.filter(
        order__in=orders
    ).values('product_name').annotate(
        total_qty=Sum('quantity'),
        total_revenue=Sum('subtotal')
    ).order_by('-total_revenue', 'product_name')[:10]

    # Sales by category — groups by the category_name SNAPSHOT stored on each
    # OrderItem at the time the order was placed (migration 0012).
    #
    # Why snapshot instead of product__category__name (live join)?
    #   1. If a product is later deleted (product FK → NULL), the live join
    #      returns NULL and those items are silently dropped from the totals.
    #   2. If a product is moved to a different category after the order,
    #      the live join reclassifies the historical sale to the new category,
    #      making the old category lose revenue it actually earned.
    #
    # Items with a blank category_name are excluded so they don't produce
    # an unnamed row in the report.
    category_sales = []
    for row in (
        OrderItem.objects
        .filter(order__in=orders)
        .exclude(category_name='')
        .values('category_name')
        .annotate(total=Sum('subtotal'), qty=Sum('quantity'))
        .order_by('-total')
    ):
        category_sales.append({
            'name':  row['category_name'],
            'total': float(row['total']),
            'qty':   row['qty'] or 0,
        })

    context = {
        'start_date':   start_date,
        'end_date':     end_date,
        'total_revenue': total_revenue,
        'total_orders':  total_orders,
        'avg_order':     avg_order,
        'daily_sales':   daily_sales,
        # True when at least one day in the range has a completed paid order.
        # Used by the template to show a "No sales in this period" placeholder
        # instead of a silently empty Daily Breakdown table/cards — the
        # {% empty %} block on the for loop cannot serve this purpose because
        # daily_sales is always non-empty (zero-filled for every day).
        'has_any_sales': any(d['count'] > 0 for d in daily_sales),
        'top_products':  top_products,
        'category_sales': category_sales,
    }
    return render(request, 'reports/index.html', context)


@login_required
@admin_required
def export_excel(request):
    """Export sales report to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return HttpResponse("openpyxl not installed", status=500)

    today = timezone.localdate()
    start_date, end_date = _parse_date_range(request, today)

    orders = (
        _base_orders_qs(start_date, end_date)
        .select_related('cashier')
        .annotate(item_count=Count('items', distinct=True))
        .order_by('created_at')
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales Report"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(fill_type="solid", fgColor="4A2C2A")

    headers = ['Order #', 'Date', 'Customer', 'Type',
               'Items', 'Total', 'Payment', 'Cashier']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font  = header_font
        cell.fill  = header_fill
        cell.alignment = Alignment(horizontal='center')

    for row, order in enumerate(orders, 2):
        ws.cell(row=row, column=1, value=order.order_number)
        # Convert stored UTC timestamp to Asia/Manila (PHT) before formatting
        # so the Excel "Date" column agrees with the date shown in the filter.
        pht_created = timezone.localtime(order.created_at)
        ws.cell(row=row, column=2, value=pht_created.strftime('%Y-%m-%d %H:%M'))
        ws.cell(row=row, column=3, value=order.customer_name)
        ws.cell(row=row, column=4, value=order.get_order_type_display())
        ws.cell(row=row, column=5, value=order.item_count)
        # Pass Decimal directly — openpyxl stores it as an exact number.
        # float(order.total) introduces IEEE 754 imprecision that can corrupt
        # Excel SUM formulas for centavo amounts like ₱123.10.
        ws.cell(row=row, column=6, value=order.total)
        ws.cell(row=row, column=7, value=order.get_payment_method_display())
        ws.cell(row=row, column=8,
                value=str(order.cashier) if order.cashier else 'N/A')

    # Auto-fit columns
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 4

    response = HttpResponse(
        content_type=(
            'application/vnd.openxmlformats-officedocument'
            '.spreadsheetml.sheet'
        )
    )
    response['Content-Disposition'] = (
        f'attachment; filename="sales_report_{start_date}_to_{end_date}.xlsx"'
    )
    wb.save(response)
    return response
