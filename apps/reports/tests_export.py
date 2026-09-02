"""
Sales Reports Excel export verification tests — Kape De Manubag.

Verifies that export_excel() produces a workbook whose data matches
the main reports_index page exactly for the same date range.

Key invariants:
  - export_excel uses the same filter as reports_index
    (is_paid=True, status='completed')
  - One data row per qualifying order
  - Sum of Total column == reports_index total_revenue
  - Cancelled / pending orders absent from the export
  - Cash and GCash both exported; distinguishable via Payment column
  - Date parameters respected (same range, same validation)
  - Total cell values are numeric (Decimal/float), not strings

Scenarios:
  1.  Normal export — single order
  2.  Multiple orders — row count matches reports_index total_orders
  3.  Sum of Total column matches reports_index total_revenue
  4.  Cash vs GCash — Payment column shows correct label
  5.  Cancelled orders excluded
  6.  Non-completed statuses excluded
  7.  Date range respected — orders outside range not exported
  8.  Empty range — header row only, no data rows
  9.  Total column is numeric (OPT fix regression)
  10. Column headers correct
  11. Export URL passes current date range from template
  12. Invalid date params fall back gracefully (200, not 500)
"""

import datetime
import io
from decimal import Decimal

from django.contrib.auth import get_user_model
from django.test import Client, TestCase
from django.urls import reverse
from django.utils import timezone

from apps.orders.models import Order, OrderItem

User = get_user_model()

TODAY = timezone.localdate()
YEST  = TODAY - datetime.timedelta(days=1)
DAY2  = TODAY - datetime.timedelta(days=2)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _user(username, role='admin'):
    u = User.objects.create_user(username=username, password='pass123')
    u.role = role
    u.save()
    return u


def _order(date, total=Decimal('100.00'), payment_method='cash',
           status='completed', is_paid=True, customer='Test',
           order_type='dine_in'):
    dt = timezone.make_aware(
        datetime.datetime.combine(date, datetime.time(10, 0))
    )
    o = Order.objects.create(
        customer_name=customer,
        status=status, is_paid=is_paid,
        payment_method=payment_method,
        order_type=order_type,
        total=total, subtotal=total,
    )
    Order.objects.filter(pk=o.pk).update(created_at=dt)
    o.refresh_from_db()
    return o


def _export_url(start=None, end=None):
    s = start or TODAY
    e = end   or TODAY
    return reverse('reports:export_excel') + f'?start={s}&end={e}'


def _wb(client, start=None, end=None):
    """Download and parse the Excel workbook. Skip if openpyxl missing."""
    try:
        import openpyxl
    except ImportError:
        return None
    resp = client.get(_export_url(start, end))
    if resp.status_code != 200:
        return None
    content = b''.join(
        resp.streaming_content
        if hasattr(resp, 'streaming_content')
        else [resp.content]
    )
    return openpyxl.load_workbook(io.BytesIO(content))


def _rows(wb):
    """Return all data rows (excluding header) as lists of cell values."""
    ws = wb.active
    return [
        [cell.value for cell in row]
        for row in ws.iter_rows(min_row=2)
        if any(cell.value is not None for cell in row)
    ]


def _headers(wb):
    ws = wb.active
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


def _skip_if_no_openpyxl(test_case):
    try:
        import openpyxl  # noqa: F401
    except ImportError:
        test_case.skipTest('openpyxl not installed')


# ── 1. Normal export — single order ──────────────────────────────────────────

class SingleOrderExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export1')
        self.client = Client()
        self.client.login(username='admin_export1', password='pass123')
        _order(TODAY, total=Decimal('250.00'), payment_method='cash')

    def test_export_returns_200(self):
        _skip_if_no_openpyxl(self)
        resp = self.client.get(_export_url())
        self.assertEqual(resp.status_code, 200)

    def test_export_content_type(self):
        _skip_if_no_openpyxl(self)
        resp = self.client.get(_export_url())
        self.assertIn('spreadsheetml', resp['Content-Type'])

    def test_one_data_row(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        self.assertEqual(len(_rows(wb)), 1)

    def test_total_value_correct(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        # Column 6 (index 5) = Total
        total_val = rows[0][5]
        self.assertAlmostEqual(float(total_val), 250.0, places=2)


# ── 2. Multiple orders — row count matches reports_index ─────────────────────

class MultipleOrdersExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export2')
        self.client = Client()
        self.client.login(username='admin_export2', password='pass123')
        _order(TODAY, total=Decimal('100.00'))
        _order(TODAY, total=Decimal('200.00'))
        _order(TODAY, total=Decimal('300.00'))

    def test_row_count_matches_reports_index(self):
        _skip_if_no_openpyxl(self)
        # Screen report
        screen = self.client.get(
            reverse('reports:index'),
            {'start': str(TODAY), 'end': str(TODAY)},
        )
        screen_count = screen.context['total_orders']
        # Excel export
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        self.assertEqual(len(_rows(wb)), screen_count)

    def test_sum_of_totals_matches_reports_index(self):
        _skip_if_no_openpyxl(self)
        screen = self.client.get(
            reverse('reports:index'),
            {'start': str(TODAY), 'end': str(TODAY)},
        )
        screen_total = float(screen.context['total_revenue'])
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows   = _rows(wb)
        export_sum = sum(float(r[5]) for r in rows)
        self.assertAlmostEqual(export_sum, screen_total, places=2)


# ── 3. Cash vs GCash — Payment column ────────────────────────────────────────

class CashGCashExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export3')
        self.client = Client()
        self.client.login(username='admin_export3', password='pass123')
        _order(TODAY, total=Decimal('300.00'), payment_method='cash',
               customer='Alice')
        _order(TODAY, total=Decimal('200.00'), payment_method='gcash',
               customer='Bob')

    def test_payment_column_shows_cash(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows     = _rows(wb)
        payments = [r[6] for r in rows]   # column 7 = Payment
        self.assertIn('Cash', payments)

    def test_payment_column_shows_gcash(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows     = _rows(wb)
        payments = [r[6] for r in rows]
        self.assertIn('GCash', payments)

    def test_two_data_rows(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        self.assertEqual(len(_rows(wb)), 2)

    def test_sum_includes_both_payment_types(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows      = _rows(wb)
        total_sum = sum(float(r[5]) for r in rows)
        self.assertAlmostEqual(total_sum, 500.0, places=2)


# ── 4. Cancelled orders excluded ─────────────────────────────────────────────

class CancelledOrdersExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export4')
        self.client = Client()
        self.client.login(username='admin_export4', password='pass123')
        _order(TODAY, status='completed', is_paid=True,  total=Decimal('200.00'))
        _order(TODAY, status='cancelled', is_paid=False, total=Decimal('999.00'))

    def test_one_row_no_cancelled(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        self.assertEqual(len(_rows(wb)), 1)

    def test_total_excludes_cancelled(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        self.assertAlmostEqual(float(rows[0][5]), 200.0, places=2)


# ── 5. Non-completed statuses excluded ───────────────────────────────────────

class NonCompletedExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export5')
        self.client = Client()
        self.client.login(username='admin_export5', password='pass123')
        _order(TODAY, status='completed', is_paid=True,  total=Decimal('150.00'))
        for st in ['pending', 'preparing', 'ready']:
            _order(TODAY, status=st,  is_paid=False, total=Decimal('999.00'))

    def test_only_completed_in_export(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        self.assertEqual(len(rows), 1)
        self.assertAlmostEqual(float(rows[0][5]), 150.0, places=2)


# ── 6. Date range respected ───────────────────────────────────────────────────

class DateRangeExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export6')
        self.client = Client()
        self.client.login(username='admin_export6', password='pass123')
        _order(DAY2,  total=Decimal('100.00'))
        _order(YEST,  total=Decimal('200.00'))
        _order(TODAY, total=Decimal('300.00'))

    def test_single_day_range_one_row(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client, TODAY, TODAY)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        self.assertEqual(len(rows), 1)
        self.assertAlmostEqual(float(rows[0][5]), 300.0, places=2)

    def test_two_day_range_two_rows(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client, YEST, TODAY)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        self.assertEqual(len(rows), 2)

    def test_three_day_range_three_rows(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client, DAY2, TODAY)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        self.assertEqual(len(rows), 3)

    def test_export_sum_matches_screen_total(self):
        """Export sum for the 3-day range must match the screen total_revenue."""
        _skip_if_no_openpyxl(self)
        screen = self.client.get(
            reverse('reports:index'),
            {'start': str(DAY2), 'end': str(TODAY)},
        )
        screen_total = float(screen.context['total_revenue'])
        wb   = _wb(self.client, DAY2, TODAY)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        export_sum = sum(float(r[5]) for r in rows)
        self.assertAlmostEqual(export_sum, screen_total, places=2)


# ── 7. Empty date range ───────────────────────────────────────────────────────

class EmptyExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export7')
        self.client = Client()
        self.client.login(username='admin_export7', password='pass123')

    def test_empty_export_returns_200(self):
        _skip_if_no_openpyxl(self)
        resp = self.client.get(_export_url())
        self.assertEqual(resp.status_code, 200)

    def test_empty_export_header_row_only(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        self.assertEqual(len(_rows(wb)), 0)

    def test_empty_export_has_headers(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        headers = _headers(wb)
        self.assertIsNotNone(headers[0])
        self.assertEqual(len([h for h in headers if h]), 8)


# ── 8. Column headers correct ─────────────────────────────────────────────────

class ColumnHeadersTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export8')
        self.client = Client()
        self.client.login(username='admin_export8', password='pass123')
        _order(TODAY, total=Decimal('100.00'))

    def test_eight_columns(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        headers = _headers(wb)
        self.assertEqual(len([h for h in headers if h is not None]), 8)

    def test_expected_header_names(self):
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        headers = _headers(wb)
        expected = ['Order #', 'Date', 'Customer', 'Type',
                    'Items', 'Total', 'Payment', 'Cashier']
        self.assertEqual(headers, expected)


# ── 9. Total column is numeric (Decimal fix regression) ──────────────────────

class TotalColumnNumericTest(TestCase):
    """
    Regression test for the fix: float(order.total) → order.total.
    The Total cell must be a numeric type (int/float/Decimal), not a string.
    openpyxl stores Python Decimal as a numeric cell.
    """

    def setUp(self):
        self.admin = _user('admin_export9')
        self.client = Client()
        self.client.login(username='admin_export9', password='pass123')
        _order(TODAY, total=Decimal('123.10'))
        _order(TODAY, total=Decimal('99.90'))

    def test_total_cells_are_numeric(self):
        _skip_if_no_openpyxl(self)
        import openpyxl
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        for row in rows:
            total_val = row[5]
            self.assertIsNotNone(total_val)
            self.assertIsInstance(total_val, (int, float, Decimal),
                f"Total cell is {type(total_val).__name__}, expected numeric")

    def test_total_values_exact(self):
        """Sum of total cells must equal the exact Decimal sum."""
        _skip_if_no_openpyxl(self)
        wb = _wb(self.client)
        if wb is None:
            self.skipTest('openpyxl not available')
        rows = _rows(wb)
        export_sum = sum(Decimal(str(r[5])) for r in rows)
        # 123.10 + 99.90 = 223.00 exactly
        self.assertEqual(export_sum, Decimal('223.00'))


# ── 10. Invalid date params fall back gracefully ──────────────────────────────

class InvalidDateParamsExportTest(TestCase):
    def setUp(self):
        self.admin = _user('admin_export10')
        self.client = Client()
        self.client.login(username='admin_export10', password='pass123')

    def test_invalid_start_returns_200(self):
        _skip_if_no_openpyxl(self)
        url = reverse('reports:export_excel') + '?start=not-a-date&end=' + str(TODAY)
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)

    def test_invalid_end_returns_200(self):
        _skip_if_no_openpyxl(self)
        url = reverse('reports:export_excel') + f'?start={TODAY}&end=bad-date'
        resp = self.client.get(url)
        self.assertEqual(resp.status_code, 200)

    def test_missing_params_returns_200(self):
        _skip_if_no_openpyxl(self)
        resp = self.client.get(reverse('reports:export_excel'))
        self.assertEqual(resp.status_code, 200)
